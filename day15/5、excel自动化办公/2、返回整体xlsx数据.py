#  xlsx   xls
# openpyxl  ->  xlsx

from openpyxl.reader.excel import load_workbook

def readXlsxFile(path):
    dic = {}
    file = load_workbook(filename=path)
    sheets = file.get_sheet_names()
    print(len(sheets))

    for sheetName in sheets:
        sheet = file.get_sheet_by_name(sheetName)
        #一张表的所有数据
        sheetInfo = []
        for lineNum in range(1, sheet.max_row + 1):
            lineList = []
            for columnNum in range(1, sheet.max_column + 1):
                value = sheet.cell(row=lineNum, column=columnNum).value
                lineList.append(value)
            sheetInfo.append(lineList)

        #将一张表的数据存到字典
        dic[sheetName] = sheetInfo
    return dic

#不能处理xls文件
path = r"C:\Users\xlg\Desktop\Python-1704\day15\5、excel自动化办公\sunck.xlsx"
dic = readXlsxFile(path)
print(dic["安力博发"])
print(len(dic))



